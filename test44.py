import tkinter as tk
import random
import time
import openpyxl
from tkinter import ttk
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from tkinter import messagebox

import openpyxl

# Workbook 객체 생성
netflix_workbook = openpyxl.Workbook()
watcha_workbook = openpyxl.Workbook()
tving_workbook = openpyxl.Workbook()

netflix_sheet = netflix_workbook.active
watcha_sheet = watcha_workbook.active
tving_sheet = tving_workbook.active

netflix_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%EB%84%B7%ED%94%8C%EB%A6%AD%EC%8A%A4+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%EB%84%B7%ED%94%8C%EB%A6%AD%EC%8A%A4+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75lsqo1LwssPveaXlssssstLC-061196"
watcha_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%EC%99%93%EC%B1%A0+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%EC%99%93%EC%B1%A0+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75dsqptbNssnJ3V2wssssss8K-371550"
tving_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%ED%8B%B0%EB%B9%99+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%ED%8B%B0%EB%B9%99+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75klqo1e8ssO4UwQhssssssud-170148"

workbooks = [netflix_workbook, watcha_workbook, tving_workbook]
sheets = [netflix_sheet, watcha_sheet, tving_sheet]
sheet_names = ["Netflix", "Watcha", "Tving"]
headers = ["제목", "국가", "장르", "평점"]
urls = [netflix_url, watcha_url, tving_url]

for sheet, sheet_name in zip(sheets, sheet_names):
    sheet.title = sheet_name
    for col, header in enumerate(headers, 1):
        sheet.cell(1, col, header)      ## 헤더 생성

def crawl_movies(country, genre):
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True) 
    #chrome_options.add_argument("--headless")          ## 백그라운드 실행
    driver = webdriver.Chrome(options=chrome_options)

    for sheet, url in zip(sheets, urls) :
        driver.get(url)
        time.sleep(0.1)

        # 국가 선택
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger2 > a > span.menu._text").click()
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger2 > div > div > div > div > div > ul > li:nth-child(" + str(country) + ") > a").click()
        

        # 장르 선택
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger3 > a > span.menu._text").click()
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger3 > div > div > div > div > div > ul > li:nth-child(" + str(genre) + ") > a").click()
        time.sleep(0.5)

        #movies = []     ## 크롤링한 영화 제목 + 평점 넣는 리스트
        n = 2
        while True:
            tag = driver.find_element(By.CLASS_NAME, '_panel')
            all_movies = tag.find_elements(By.CLASS_NAME, 'info_box')
            for movie in all_movies: 
                names = movie.find_element(By.CLASS_NAME, "title")
                MovieName = names.find_element(By.TAG_NAME, "a").text
                now_country = driver.find_element(By.XPATH, "//*[@id=\"main_pack\"]/section[1]/div[2]/div/div/div[1]/div/div/ul/li[2]/a").text
                now_genre = driver.find_element(By.XPATH, "//*[@id=\"main_pack\"]/section[1]/div[2]/div/div/div[1]/div/div/ul/li[3]/a").text
                try:
                    rating = float(movie.find_element(By.CLASS_NAME, "num").text)
                except:
                    rating = 0.0
                if MovieName == '':
                    continue
                #movies.append([MovieName, rating])
                sheet.cell(n,1,MovieName)
                sheet.cell(n,2,now_country)
                sheet.cell(n,3,now_genre)
                sheet.cell(n,4,rating)
                n += 1
                
            try:        ## 다음창 넘기는 부분
                next = driver.find_element( By.XPATH, '//*[@id="main_pack"]/section[1]/div[2]/div/div/div[3]/div/a[2]')
                nowpage = driver.find_element(By.CLASS_NAME, 'npgs_now._current').text
                nextpage = driver.find_element(By.CLASS_NAME, '_total').text
                next.click()
                time.sleep(0.1)
                if int(nowpage) == int(nextpage):
                        break
            except :    
                break
            
        if sheet == netflix_sheet :
            netflix_workbook.save("C:\\CookAnalysis\\Excel\\Netflix.xlsx")
        elif sheet == watcha_sheet :
            watcha_workbook.save("C:\\CookAnalysis\\Excel\\Watcha.xlsx")
        else :
            tving_workbook.save("C:\\CookAnalysis\\Excel\\Tving.xlsx")
        
    print("파일 저장 완료")
    driver.quit()

    
    #if len(movies) > 2:
    #    movies.sort(key=lambda x: x[1], reverse=True)


    # tkinter 윈도우 생성
    ''' 
    if len(movies) > 0 :
        root = tk.Tk()
        root.title("영화 평점 순위")
        tree = ttk.Treeview(root)
        tree["columns"] = ("영화 제목", "평점")
        tree.column("#0", width=0, stretch=tk.NO)
        tree.column("영화 제목", anchor=tk.CENTER, width=300)
        tree.column("평점", anchor=tk.CENTER, width=100)
        tree.heading("#0", text="", anchor=tk.CENTER)
        tree.heading("영화 제목", text="영화 제목", anchor=tk.CENTER)
        tree.heading("평점", text="평점", anchor=tk.CENTER)

    # 트리뷰에 데이터 삽입
        for idx, (movie, rating) in enumerate(movies, start=1):
            tree.insert("", "end", text=str(idx), values=(movie, rating))

    # 스크롤바 생성
        vsb = ttk.Scrollbar(root, orient="vertical", command=tree.yview)
        vsb.pack(side='right', fill='y')
        tree.configure(yscrollcommand=vsb.set)

    # 트리뷰 배치
        tree.pack(expand=True, fill='both')

        driver.quit()
        root.mainloop()
    '''


def on_submit():
    # 콤보박스에서 선택한 인덱스를 가져옴
    selected_country = country_combo.current() + 1  # 인덱스는 0부터 시작하므로 +1
    selected_genre = genre_combo.current() + 1      # 인덱스는 0부터 시작하므로 +1
    crawl_movies(selected_country, selected_genre)


country = ["전체", "한국", "미국", "일본", "중국", "대만", "영국", "해외"]
genre = ["전체", "공포", "스릴러", "로맨스", "코미디", "로맨틱코미디", "액션", "하이틴", "가족", "어린이", "SF", "판타지", "수사", "좀비", "재난", "전쟁", "에로",
         "BL", "추리", "범죄", "반전", "학교", "시대극", "우주", "시트콤", "음악", "무협", "슬픈", "감동적인", "힐링되는", "명작고전", "실화바탕", "웹툰원작", "마블", "디즈니", "연애"]

# Tkinter 윈도우 생성
window = tk.Tk()
window.title("네이버 영화 크롤링 및 엑셀 파일 열기")

# 국가 선택 콤보박스
country_label = tk.Label(window, text="국가 선택:")
country_label.grid(row=0, column=0, padx=10, pady=5)
country_combo = ttk.Combobox(window, values=country)
country_combo.grid(row=0, column=1, padx=10, pady=5)
country_combo.current(0)

# 장르 선택 콤보박스
genre_label = tk.Label(window, text="장르 선택:")
genre_label.grid(row=1, column=0, padx=10, pady=5)
genre_combo = ttk.Combobox(window, values=genre)
genre_combo.grid(row=1, column=1, padx=10, pady=5)
genre_combo.current(0)

# 선택완료 버튼
submit_button = tk.Button(window, text="선택완료", command=on_submit)
submit_button.grid(row=2, column=0, columnspan=2, padx=10, pady=10)

window.mainloop()
