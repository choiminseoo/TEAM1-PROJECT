import tkinter as tk
import random
import time
import openpyxl
import bs4
import urllib.request
from tkinter import ttk
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from tkinter import messagebox
from lxml import etree
from openpyxl.styles import PatternFill

netflix_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%EB%84%B7%ED%94%8C%EB%A6%AD%EC%8A%A4+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%EB%84%B7%ED%94%8C%EB%A6%AD%EC%8A%A4+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75lsqo1LwssPveaXlssssstLC-061196"
watcha_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%EC%99%93%EC%B1%A0+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%EC%99%93%EC%B1%A0+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75dsqptbNssnJ3V2wssssss8K-371550"
tving_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%ED%8B%B0%EB%B9%99+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%ED%8B%B0%EB%B9%99+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75klqo1e8ssO4UwQhssssssud-170148"
urls = [netflix_url, watcha_url, tving_url]
headers = ["제목", "국가", "장르", "평점", "개봉연도", "해시태그", "OTT"]

def crawl_movies(country, genre):
    rawData_workbook = openpyxl.Workbook()
    rawData_sheet = rawData_workbook.active
    rawData_sheet.title = "RawData"
    for col, header in enumerate(headers, 1):
        rawData_sheet.cell(1, col, header)
        rawData_sheet.cell(1, col).fill = PatternFill(start_color = "B7F0B1", fill_type= "solid")

    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True) 
    #chrome_options.add_argument("--headless")          ## 백그라운드 실행
    driver = webdriver.Chrome(options=chrome_options)

    for url in urls :
        driver.get(url)
        time.sleep(0.1)

        # 국가 선택
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger2 > a > span.menu._text").click()
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger2 > div > div > div > div > div > ul > li:nth-child(" + str(country) + ") > a").click()
        
        # 장르 선택
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger3 > a > span.menu._text").click()
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger3 > div > div > div > div > div > ul > li:nth-child(" + str(genre) + ") > a").click()
        time.sleep(0.5)

        movieCountry = driver.find_element(By.XPATH, "//*[@id=\"main_pack\"]/section[1]/div[2]/div/div/div[1]/div/div/ul/li[2]/a").text
        movieGenre = driver.find_element(By.XPATH, "//*[@id=\"main_pack\"]/section[1]/div[2]/div/div/div[1]/div/div/ul/li[3]/a").text

        pageIdx = 1
        while True :
            for i in range(1, 9) :
                try :
                    movieName = driver.find_element(By.XPATH, "//*[@id=\"mflick\"]/div/div/ul["+str(pageIdx)+"]/li["+str(i)+"]/strong/a").text
                except :
                    continue
        
                try :
                    a = driver.find_element(By.XPATH, "//*[@id=\"mflick\"]/div/div/ul["+str(pageIdx)+"]/li["+str(i)+"]/div[2]/span[3]/span").text
                    if "." in a :
                        avgScore = a
                    else :
                        avgScore = "평점정보X"
                except :
                    avgScore = "평점정보X"

                try :
                    y = driver.find_element(By.XPATH, "//*[@id=\"mflick\"]/div/div/ul["+str(pageIdx)+"]/li["+str(i)+"]/div[2]/span[2]").text
                    if len(y) == 4 :
                        year = y
                    else :
                        year = "년도정보X"
                except :
                    year = "년도정보X"

                try :
                    h = driver.find_element(By.XPATH, "//*[@id=\"mflick\"]/div/div/ul["+str(pageIdx)+"]/li["+str(i)+"]/div[3]").text
                    if "#" in h :
                        hashTag = h
                    else :
                        hashTag = "태그정보X"
                except :
                    hashTag = "태그정보X"

                movieInfo = [movieName, movieCountry, movieGenre, avgScore, year, hashTag]

                if url == netflix_url :
                    movieInfo.append("Netflix")

                elif url == watcha_url :
                    movieInfo.append("Watcha")

                else :
                    movieInfo.append("Tving")

                print(movieInfo)
                rawData_sheet.append(movieInfo)
                time.sleep(0.1)

            try: 
                next = driver.find_element( By.XPATH, '//*[@id="main_pack"]/section[1]/div[2]/div/div/div[3]/div/a[2]')
                nowpage = driver.find_element(By.CLASS_NAME, 'npgs_now._current').text
                nextpage = driver.find_element(By.CLASS_NAME, '_total').text
                next.click()
                pageIdx += 1
                if int(nowpage) == int(nextpage):
                    break
            except :    
                break
    
    rawData_workbook.save("C:\\CookAnalysis\\Excel\\RowData.xlsx")
    driver.quit()

def on_submit():
    # 콤보박스에서 선택한 인덱스를 가져옴
    selected_country = country_combo.current() + 2  # 인덱스는 0부터 시작하므로 +1
    selected_genre = genre_combo.current() + 2      # 인덱스는 0부터 시작하므로 +1
    crawl_movies(selected_country, selected_genre)


country = ["한국", "미국", "일본", "중국", "대만", "영국", "해외"]
genre = ["공포", "스릴러", "로맨스", "코미디", "로맨틱코미디", "액션", "하이틴", "가족", "어린이", "SF", "판타지", "수사", "좀비", "재난", "전쟁", "에로",
         "BL", "추리", "범죄", "반전", "학교", "시대극", "우주", "시트콤", "음악", "무협", "슬픈", "감동적인", "힐링되는", "명작고전", "실화바탕", "웹툰원작", "마블", "디즈니", "연애"]

# Tkinter 윈도우 생성
window = tk.Tk()
window.title("네이버 영화 크롤링 및 엑셀 파일 열기")

# 국가 선택 콤보박스
country_label = tk.Label(window, text="국가 선택:")
country_label.grid(row=0, column=0, padx=10, pady=5)
country_combo = ttk.Combobox(window, values=country)
country_combo.grid(row=0, column=1, padx=10, pady=5)
country_combo.current(0)

# 장르 선택 콤보박스
genre_label = tk.Label(window, text="장르 선택:")
genre_label.grid(row=1, column=0, padx=10, pady=5)
genre_combo = ttk.Combobox(window, values=genre)
genre_combo.grid(row=1, column=1, padx=10, pady=5)
genre_combo.current(0)

# 선택완료 버튼
submit_button = tk.Button(window, text="선택완료", command=on_submit)
submit_button.grid(row=2, column=0, columnspan=2, padx=10, pady=10)

window.mainloop()