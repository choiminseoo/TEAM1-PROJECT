import tkinter as tk
import openpyxl
import tkinter.messagebox as msgbox
from tkinter import ttk
from openpyxl.styles import PatternFill

def login():
    try :
        workbook = openpyxl.load_workbook("C:\\CookAnalysis\\Excel\\userData.xlsx")
        worksheet = workbook["userData"]
        username = username_entry.get()
        password = password_entry.get()
        if (username == worksheet.cell(2, 1).value and password == worksheet.cell(2,2).value) :
            msgbox.showinfo("알림", "로그인 성공 !")
        elif (username != worksheet.cell(2, 1).value and password == worksheet.cell(2,2).value) :
            msgbox.showinfo("ERROR", "아이디가 맞지 않습니다.")
        elif (username == worksheet.cell(2, 1).value and password != worksheet.cell(2,2).value) :
            msgbox.showinfo("ERROR", "비밀번호가 맞지 않습니다.")
        else :
            msgbox.showinfo("ERROR", "아이디와 비밀번호가 맞지 않습니다.")
    except :
        msgbox.showinfo("알림", "계정이 없으니 회원가입 먼저 하세요")



def register_window():
    global reg_username_entry, reg_password_entry, reg_name_entry, reg_age_entry, reg_likeGenre_combo, reg_haveOtt_combo, gender_var, register_windows

    register_windows = tk.Toplevel(root)
    register_windows.title("회원가입")

    # 아이디 입력 레이블 및 엔트리 위젯
    reg_username_label = tk.Label(register_windows, text="아이디:")
    reg_username_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
    reg_username_entry = tk.Entry(register_windows)
    reg_username_entry.grid(row=0, column=1, padx=5, pady=5)


    # 비밀번호 입력 레이블 및 엔트리 위젯
    reg_password_label = tk.Label(register_windows, text="비밀번호:")
    reg_password_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
    reg_password_entry = tk.Entry(register_windows, show="*")  # 입력 내용을 마스킹합니다.
    reg_password_entry.grid(row=1, column=1, padx=5, pady=5)


    reg_name_label = tk.Label(register_windows, text="이름:")
    reg_name_label.grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
    reg_name_entry = tk.Entry(register_windows)
    reg_name_entry.grid(row=2, column=1, padx=5, pady=5)

    gender_label = tk.Label(register_windows, text="성별:")
    gender_label.grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)
    gender_var = tk.StringVar(value="남자")
    male_radio = tk.Radiobutton(register_windows, text="남자", variable=gender_var, value="남자")
    male_radio.grid(row=3, column=1, padx=5, pady=5, sticky=tk.W)
    female_radio = tk.Radiobutton(register_windows, text="여자", variable=gender_var, value="여자")
    female_radio.grid(row=3, column=2, padx=5, pady=5, sticky=tk.W)

    reg_age_label = tk.Label(register_windows, text="나이:")
    reg_age_label.grid(row=4, column=0, padx=5, pady=5, sticky=tk.W)
    reg_age_entry = tk.Entry(register_windows)
    reg_age_entry.grid(row=4, column=1, padx=5, pady=5)


    likeGenre = ["공포", "스릴러", "로맨스", "코미디", "로맨틱코미디", "액션", "하이틴", "가족", "어린이", "SF", "판타지", "수사", "좀비", "재난", "전쟁", "에로",
                 "BL", "추리", "범죄", "반전", "학교", "시대극", "우주", "시트콤", "음악", "무협", "슬픈", "감동적인", "힐링되는", "명작고전", "실화바탕", "웹툰원작", "마블", "디즈니",
                 "연애"]
    reg_likeGenre_label = tk.Label(register_windows, text="좋아하는 장르:")
    reg_likeGenre_label.grid(row=5, column=0, padx=5, pady=5, sticky=tk.W)
    reg_likeGenre_combo = ttk.Combobox(register_windows, values=likeGenre)
    reg_likeGenre_combo.grid(row=5, column=1, padx=5, pady=5)
    reg_likeGenre_combo.current(0)

    haveOtt = ["없음", "넷플릭스", "왓챠", "티빙"]
    reg_haveOtt_label = tk.Label(register_windows, text="보유한 OTT:")
    reg_haveOtt_label.grid(row=6, column=0, padx=5, pady=5, sticky=tk.W)
    reg_haveOtt_combo = ttk.Combobox(register_windows, values=haveOtt)
    reg_haveOtt_combo.grid(row=6, column=1, padx=5, pady=5)
    reg_haveOtt_combo.current(0)


    # 입력완료 버튼
    register_button2 = tk.Button(register_windows, text="입력완료", command= register)
    register_button2.grid(row=7, column=0, columnspan=2, pady=10)

    

def register():
    global userInfo

    headers = ["아이디", "비밀번호", "이름", "성별", "나이", "좋아하는 장르", "보유OTT"]
    userData_workbook = openpyxl.Workbook()
    userData_sheet = userData_workbook.active
    userData_sheet.title = "userData"
    for col, header in enumerate(headers, 1):
        userData_sheet.cell(1, col, header)
        userData_sheet.cell(1, col).fill = PatternFill(start_color = "B7F0B1", fill_type= "solid")

    reUsername = reg_username_entry.get()
    rePassword = reg_password_entry.get()
    reName = reg_name_entry.get()
    reAge = reg_age_entry.get()
    reGenre = reg_likeGenre_combo.get()
    reOTT = reg_haveOtt_combo.get()
    genderValue = gender_var.get()
    userInfo = [reUsername, rePassword, reName, genderValue, reAge, reGenre, reOTT]
    userData_sheet.append(userInfo)
    userData_workbook.save("C:\\CookAnalysis\\Excel\\userData.xlsx")
    msgbox.showinfo("알림", "회원가입 완료") 
    register_windows.destroy()



# Tkinter 창 생성
root = tk.Tk()
root.geometry("230x160")
root.title("로그인")

# 아이디 입력 레이블 및 엔트리 위젯
username_label = tk.Label(root, text="아이디:")
username_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
username_entry = tk.Entry(root)
username_entry.grid(row=0, column=1, padx=5, pady=5)

# 비밀번호 입력 레이블 및 엔트리 위젯
password_label = tk.Label(root, text="비밀번호:")
password_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
password_entry = tk.Entry(root, show="*")  # 입력 내용을 마스킹합니다.
password_entry.grid(row=1, column=1, padx=5, pady=5)

# 로그인, 회원가입 버튼
login_button = tk.Button(root, text="로그인", command=login)
login_button.grid(row=2, column=0, columnspan=2, pady=10)
register_button = tk.Button(root, text="회원가입", command=register_window)
register_button.grid(row=3, column=0, columnspan=2, pady=10)

# 창 실행
root.mainloop()