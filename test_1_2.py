import tkinter as tk
import time
import openpyxl
import bs4
import urllib.request
from tkinter import ttk
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from tkinter import messagebox
from lxml import etree
from openpyxl.styles import PatternFill

## 헤더(유저)정보
#header01 = {"User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"}

## Workbook 객체 생성
rawData_workbook = openpyxl.Workbook()
netflix_workbook = openpyxl.Workbook()
watcha_workbook = openpyxl.Workbook()
tving_workbook = openpyxl.Workbook()

## 시트 활성화
rawData_sheet = rawData_workbook.active
netflix_sheet = netflix_workbook.active
watcha_sheet = watcha_workbook.active
tving_sheet = tving_workbook.active

## 영화 목록 url
netflix_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%EB%84%B7%ED%94%8C%EB%A6%AD%EC%8A%A4+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%EB%84%B7%ED%94%8C%EB%A6%AD%EC%8A%A4+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75lsqo1LwssPveaXlssssstLC-061196"
watcha_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%EC%99%93%EC%B1%A0+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%EC%99%93%EC%B1%A0+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75dsqptbNssnJ3V2wssssss8K-371550"
tving_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%ED%8B%B0%EB%B9%99+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%ED%8B%B0%EB%B9%99+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75klqo1e8ssO4UwQhssssssud-170148"

workbooks = [netflix_workbook, watcha_workbook, tving_workbook]     ## 워크북 모음
sheets = [rawData_sheet, netflix_sheet, watcha_sheet, tving_sheet]      ## 시트 모음
sheet_names = ["RawData", "Netflix", "Watcha", "Tving"]     ## 시트 이름 모음
headers = ["제목", "국가", "장르", "평점", "러닝타임", "개봉일", "관객수", "감독", "출연", "OTT"]   ## 헤더 모음
urls = [netflix_url, watcha_url, tving_url]     ## url 모음

## 시트에 타이틀 부여, 헤더 생성, 헤더 배경색 생성
for sheet, sheet_name in zip(sheets, sheet_names):
    sheet.title = sheet_name
    for col, header in enumerate(headers, 1):
        sheet.cell(1, col, header)
        sheet.cell(1, col).fill = PatternFill(start_color = "B7F0B1", fill_type= "solid")


## 영화 정보 크롤링 함수
def crawl_movies(country, genre):
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True) 
    #chrome_options.add_argument("--headless")          ## 백그라운드 실행
    driver = webdriver.Chrome(options=chrome_options)
    
    ## 3번(넷플릭스 url, 왓챠 url, 티빙 url) 반복 
    for url in urls :
        driver.get(url)
        time.sleep(0.1)

        # 국가 선택
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger2 > a > span.menu._text").click()
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger2 > div > div > div > div > div > ul > li:nth-child(" + str(country) + ") > a").click()
        
        # 장르 선택
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger3 > a > span.menu._text").click()
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger3 > div > div > div > div > div > ul > li:nth-child(" + str(genre) + ") > a").click()
        time.sleep(0.5)

        pageIdx = 1   ## 영화 제목 xpath의 페이지 값

        while True :
            ## 한 페이지 최대 영화 개수 만큼 반복
            for i in range(1, 9) :
                try :    ## 셀레니움으로 영화 제목 가져온 후 리퀘스트로 영화 정보 가져오기
                    movieName = driver.find_element(By.XPATH, "//*[@id=\"mflick\"]/div/div/ul["+str(pageIdx)+"]/li["+str(i)+"]/strong/a").text
                    infoUrl = "https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=0&ie=utf8&query=" + urllib.parse.quote("영화" + movieName)
                    htmlObject = urllib.request.urlopen(infoUrl)
                    webPage = htmlObject.read()
                    bsObject = bs4.BeautifulSoup(webPage, 'html.parser')
                    html_tree = etree.HTML(webPage)

                    movieOverview = html_tree.xpath("//*[@id=\"main_pack\"]/div[2]/div[2]/div[2]/div/div[1]/dl/div[1]/dd/text()")
                    if movieOverview == [] :
                        movieOverview = html_tree.xpath("//*[@id=\"main_pack\"]/div[2]/div[2]/div[1]/div/div[1]/dl/div[1]/dd/text()")
                    
                    movieGenre = movieOverview[0]
                    movieCountry = movieOverview[1]

                    ## 러닝타임이 없는 경우
                    try :
                        runningTime = movieOverview[2]
                    except :
                        runningTime = "시간정보X"

                    ## 평점 정보
                    try :
                        avgScore = html_tree.xpath("//*[@id=\"main_pack\"]/div[2]/div[2]/div[2]/div/div[1]/dl/div[3]/dd/text()")[0]
                    except :
                        avgScore = "평점정보X"
                    
                    ## 개봉일 정보
                    try :
                        releaseDate = html_tree.xpath("//*[@id=\"main_pack\"]/div[2]/div[2]/div[2]/div/div[1]/dl/div[2]/dd/text()")[0]
                    except :
                        releaseDate = "개봉정보X"
                    
                    ## 괸객수 정보
                    try :
                        audience = html_tree.xpath("//*[@id=\"main_pack\"]/div[2]/div[2]/div[2]/div/div[1]/dl/div[4]/dd/text()")[0]
                    except :
                        audience = "관객정보X"
                    
                    ## 감독, 배우 정보
                    director = None
                    actors = ""
                    tag = bsObject.find("div", {"class" : "list_info _scroller"})
                    strongTag = tag.find_all("strong", {"class": "name type_ell_2 _html_ellipsis"})

                    try :
                        for n, tag_text in enumerate(strongTag, 1) :
                            if n == 1 :
                                director = tag_text.text
                            else :
                                actors += tag_text.text
                                if n != len(strongTag) :
                                    actors += ", "
                    except :
                        director = "감독정보X"
                        actors = "배우정보X"

                    print(movieName, movieCountry, movieGenre, avgScore, runningTime, releaseDate, audience, director, actors)
                    movieInfo = [movieName, movieCountry, movieGenre, avgScore, runningTime, releaseDate, audience, director, actors]
                    
                    if url == netflix_url :
                        movieInfo.append("Netflix")

                    elif url == watcha_url :
                        movieInfo.append("Watcha")

                    else :
                        movieInfo.append("Tving")
                    
                    rawData_sheet.append(movieInfo)
                    time.sleep(0.1)
                
                except :
                    continue
            
            ## 다음 페이지 넘기기
            try: 
                next = driver.find_element( By.XPATH, '//*[@id="main_pack"]/section[1]/div[2]/div/div/div[3]/div/a[2]')
                nowpage = driver.find_element(By.CLASS_NAME, 'npgs_now._current').text
                nextpage = driver.find_element(By.CLASS_NAME, '_total').text
                next.click()
                pageIdx += 1
                if int(nowpage) == int(nextpage):
                    break
            except :    
                    break
        
        rawData_workbook.save("C:\\CookAnalysis\\Excel\\RowData.xlsx")


    driver.quit()

    ## RowData 파일 열기
    workbook = openpyxl.load_workbook("C:\\CookAnalysis\\Excel\\RowData.xlsx")
    data_sheet = workbook.active
    
    ## RowData에서 OTT 별로 정보 추출 후, 각 OTT 파일에 담기
    for row in data_sheet.iter_rows(min_row = 2) :
        if row[9].value == "Netflix" :    ## OTT 정보 인덱스 값은 9
            copyRow = []
            for idx in range(10) : 
                copyRow.append(row[idx].value)
            netflix_sheet.append(copyRow)
        elif row[9].value == "Watcha" :
            copyRow = []
            for idx in range(10) :
                copyRow.append(row[idx].value)
            watcha_sheet.append(copyRow)
        else :
            copyRow = []
            for idx in range(10) :
                copyRow.append(row[idx].value)
            tving_sheet.append(copyRow)

    netflix_workbook.save("C:\\CookAnalysis\\Excel\\Netflix.xlsx")
    watcha_workbook.save("C:\\CookAnalysis\\Excel\\Watcha.xlsx")
    tving_workbook.save("C:\\CookAnalysis\\Excel\\Tving.xlsx")
    print("정리 완료")


def on_submit():
    # 콤보박스에서 선택한 인덱스를 가져옴
    selected_country = country_combo.current() + 1  # 인덱스는 0부터 시작하므로 +1
    selected_genre = genre_combo.current() + 1      # 인덱스는 0부터 시작하므로 +1
    crawl_movies(selected_country, selected_genre)


country = ["전체", "한국", "미국", "일본", "중국", "대만", "영국", "해외"]
genre = ["전체", "공포", "스릴러", "로맨스", "코미디", "로맨틱코미디", "액션", "하이틴", "가족", "어린이", "SF", "판타지", "수사", "좀비", "재난", "전쟁", "에로",
         "BL", "추리", "범죄", "반전", "학교", "시대극", "우주", "시트콤", "음악", "무협", "슬픈", "감동적인", "힐링되는", "명작고전", "실화바탕", "웹툰원작", "마블", "디즈니", "연애"]

# Tkinter 윈도우 생성
window = tk.Tk()
window.title("네이버 영화 크롤링 및 엑셀 파일 열기")

# 국가 선택 콤보박스
country_label = tk.Label(window, text="국가 선택:")
country_label.grid(row=0, column=0, padx=10, pady=5)
country_combo = ttk.Combobox(window, values=country)
country_combo.grid(row=0, column=1, padx=10, pady=5)
country_combo.current(0)

# 장르 선택 콤보박스
genre_label = tk.Label(window, text="장르 선택:")
genre_label.grid(row=1, column=0, padx=10, pady=5)
genre_combo = ttk.Combobox(window, values=genre)
genre_combo.grid(row=1, column=1, padx=10, pady=5)
genre_combo.current(0)

# 선택완료 버튼
submit_button = tk.Button(window, text="선택완료", command=on_submit)
submit_button.grid(row=2, column=0, columnspan=2, padx=10, pady=10)

window.mainloop()